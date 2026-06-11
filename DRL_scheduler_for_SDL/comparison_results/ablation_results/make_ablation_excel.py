#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Aggregate ERC ablation result JSONs into a wide per-case comparison Excel.

Usage:
    python make_ablation_excel.py [results_dir] [output.xlsx]

Both arguments are optional:
    results_dir   directory containing episode_*_seed_*.json   (default: this script's folder)
    output.xlsx   output workbook path                         (default: <results_dir>/ablation_summary.xlsx)

Generic by design:
    * Auto-discovers every  episode_<idx>_seed_<seed>.json  in results_dir,
      so the number of result files and their seeds are irrelevant.
    * Control groups are auto-discovered from the JSON content (any top-level
      key that looks like an ablation arm, i.e. carries a 'makespan' field),
      so adding/removing arms does not break the script. A preferred display
      order is applied when known groups are present, with any extra groups
      appended afterwards.

Layout produced:
    one row  = one case (episode/seed)
    columns  = [episode, seed] + for each group: makespan, total_solve_time,
               total_releases, total_evaluations
    headers  = two rows (group on top, metric below); each group color-coded.

Metric definitions:
    makespan            -> group['makespan']
    total_solve_time    -> group['total_solve_time']
    total_releases      -> number of entries in group['release_history']
    total_evaluations   -> sum of 'combinations_tried' across release_history
"""

import os
import re
import sys
import json
import glob
import statistics as st

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


METRICS = ['makespan', 'total_solve_time', 'total_releases', 'total_evaluations']

# Preferred display order for the standard ERC ablation arms. Any group found in
# the data but not listed here is appended (sorted) after these.
PREFERRED_GROUP_ORDER = [
    'erc_full',
    'erc_no_forward',
    'erc_no_backward',
    'erc_no_urgency',
]

# Distinct fill per group, cycled if there are more groups than colors.
FILL_PALETTE = [
    'DDEBF7',  # blue
    'E2EFDA',  # green
    'FCE4D6',  # orange
    'FFF2CC',  # yellow
    'EAD1DC',  # pink
    'D9D2E9',  # purple
    'F4CCCC',  # red
    'D0E0E3',  # teal
]

FILENAME_RE = re.compile(r'episode_(\d+)_seed_(\d+)\.json$')


def discover_groups(records):
    """Return ordered list of control-group keys present across all JSON records.

    A top-level key counts as a group if its value is a dict containing a
    'makespan' field. Order: PREFERRED_GROUP_ORDER first (those that appear),
    then any remaining groups sorted alphabetically.
    """
    found = set()
    for d in records:
        for k, v in d.items():
            if isinstance(v, dict) and 'makespan' in v:
                found.add(k)
    ordered = [g for g in PREFERRED_GROUP_ORDER if g in found]
    extras = sorted(g for g in found if g not in PREFERRED_GROUP_ORDER)
    return ordered + extras


def metric_values(group_dict):
    """Extract the four metrics from one group's result dict."""
    rh = group_dict.get('release_history', []) or []
    ms = group_dict.get('makespan')
    tst = group_dict.get('total_solve_time')
    return [
        round(ms, 3) if isinstance(ms, (int, float)) else ms,
        round(tst, 3) if isinstance(tst, (int, float)) else tst,
        len(rh),
        sum(int(e.get('combinations_tried', 0) or 0) for e in rh),
    ]


def load_cases(results_dir):
    """Load all episode_*_seed_*.json files. Returns (records, parsed_meta)."""
    files = sorted(glob.glob(os.path.join(results_dir, 'episode_*_seed_*.json')))
    records, meta = [], []
    for fp in files:
        m = FILENAME_RE.search(os.path.basename(fp))
        if not m:
            continue
        try:
            with open(fp, 'r', encoding='utf-8') as f:
                d = json.load(f)
        except (json.JSONDecodeError, OSError) as e:
            print('  [skip] %s (%s)' % (os.path.basename(fp), e))
            continue
        # Prefer explicit fields inside the JSON; fall back to the filename.
        ep = d.get('episode_idx', int(m.group(1)))
        seed = d.get('seed', int(m.group(2)))
        records.append(d)
        meta.append((ep, seed))
    return records, meta


def build_rows(records, meta, groups):
    """Build the per-case data rows (a list of flat value lists)."""
    rows = []
    for d, (ep, seed) in zip(records, meta):
        rec = [ep, seed]
        for g in groups:
            gd = d.get(g, {}) or {}
            rec += metric_values(gd)
        rows.append(rec)
    rows.sort(key=lambda r: (r[0], r[1]))
    return rows


def summary_rows(rows, groups):
    """Compute mean and std (population) for every metric column."""
    n_metric_cols = len(groups) * len(METRICS)
    mean_row = ['mean', '']
    std_row = ['std', '']
    for ci in range(2, 2 + n_metric_cols):
        vals = [r[ci] for r in rows if isinstance(r[ci], (int, float))]
        if vals:
            mean_row.append(round(st.mean(vals), 3))
            std_row.append(round(st.pstdev(vals), 3) if len(vals) > 1 else 0)
        else:
            mean_row.append(None)
            std_row.append(None)
    return [mean_row, std_row]


def build_sheet(ws, data_rows, groups, group_fills):
    """Write a two-row header + data block onto worksheet ws."""
    hdr_font = Font(bold=True)
    center = Alignment(horizontal='center', vertical='center')

    ws.cell(row=2, column=1, value='episode').font = hdr_font
    ws.cell(row=2, column=2, value='seed').font = hdr_font
    ws.cell(row=2, column=1).alignment = center
    ws.cell(row=2, column=2).alignment = center

    start = 3
    for g in groups:
        c0 = start
        c1 = start + len(METRICS) - 1
        ws.merge_cells(start_row=1, start_column=c0, end_row=1, end_column=c1)
        hc = ws.cell(row=1, column=c0, value=g)
        hc.font = hdr_font
        hc.alignment = center
        for j, met in enumerate(METRICS):
            cc = c0 + j
            ws.cell(row=1, column=cc).fill = group_fills[g]
            mc = ws.cell(row=2, column=cc, value=met)
            mc.font = hdr_font
            mc.alignment = center
            mc.fill = group_fills[g]
        start = c1 + 1

    for ri, rec in enumerate(data_rows):
        for cj, val in enumerate(rec):
            ws.cell(row=3 + ri, column=1 + cj, value=val)

    ws.column_dimensions['A'].width = 9
    ws.column_dimensions['B'].width = 13
    last_col = 2 + len(groups) * len(METRICS)
    for cc in range(3, last_col + 1):
        ws.column_dimensions[get_column_letter(cc)].width = 16
    ws.freeze_panes = 'C3'


def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    results_dir = sys.argv[1] if len(sys.argv) > 1 else script_dir
    out = sys.argv[2] if len(sys.argv) > 2 else os.path.join(results_dir, 'ablation_summary.xlsx')

    records, meta = load_cases(results_dir)
    if not records:
        print('No episode_*_seed_*.json files found in: %s' % results_dir)
        sys.exit(1)

    groups = discover_groups(records)
    if not groups:
        print('No control-group data (no dict with a "makespan" field) found.')
        sys.exit(1)

    group_fills = {
        g: PatternFill('solid', fgColor=FILL_PALETTE[i % len(FILL_PALETTE)])
        for i, g in enumerate(groups)
    }

    rows = build_rows(records, meta, groups)
    summ = summary_rows(rows, groups)

    wb = Workbook()
    ws1 = wb.active
    ws1.title = 'detail_wide'
    build_sheet(ws1, rows, groups, group_fills)

    ws2 = wb.create_sheet('summary_by_group')
    build_sheet(ws2, summ, groups, group_fills)

    wb.save(out)
    print('Cases: %d   Groups: %s' % (len(rows), ', '.join(groups)))
    print('Written: %s' % out)


if __name__ == '__main__':
    main()
